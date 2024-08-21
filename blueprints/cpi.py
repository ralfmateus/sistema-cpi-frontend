from datetime import date
from flask import Blueprint, redirect, render_template, abort, request, send_file, url_for, session
import requests
from openpyxl import Workbook
from openpyxl.styles import Border, Color, PatternFill, Font, Alignment, Side   
from jinja2 import TemplateNotFound

from blueprints.artigo import ARTIGOS
# from blueprints.usuario import verificar_funcao

cpi = Blueprint('cpi', __name__, url_prefix='/cpi')


@cpi.route('/adicionar', methods=['GET', 'POST'])
def adicionar_cpi():
    if not session.get('token'):
        return redirect(url_for('usuario.login'))
    
    headers = {'Authorization': f'Bearer {session["token"]}'}
    
    response = requests.get(f'http://127.0.0.1:8000/usuario/get_usuario', headers=headers)
    
    if not response:
        del session['token']
        return redirect(url_for('usuario.login', erro=1))
    
    usuario = response.json()['__data__']
    
    response = requests.get(f'http://127.0.0.1:8000/usuario/get_by_id?id={usuario["id"]}', headers=headers)
    
    if not response:
        del session['usuario']
        return redirect(url_for('usuario.login', erro=1))
    
    usuario = response.json()
    
    
    response = requests.get('http://127.0.0.1:8000/usuario/get_all', headers=headers)
    comunicantes = response.json()
    
    if request.method == 'GET':
        data = date.today()
        
        return render_template('adicionar_cpi.html', artigos=ARTIGOS, comunicantes=comunicantes, data=data)
    else:
        comunicante = int(request.form.get('comunicante'))
        temp = request.form.get('data')
        temp = temp.split('-')
        data = f'{temp[2]}/{temp[1]}/{temp[0]}'
        numero = int(request.form.get('numero'))
        
        curso = request.form.get('curso')
        ano = request.form.get('ano')
        
        params = {
            'numero': numero,
            'curso': curso,
            'ano': ano,
        }
    
        response = requests.get('http://127.0.0.1:8000/info/get_by_numero_curso_ano', params=params, headers=headers)
        
        info = response.json()
        
        
        cpi = {
            "aluno": info[0]['usuario']['id'],
            "conduta": request.form.get('conduta'),
            "data": data,
            "hora": request.form.get('hora'),
            "local": request.form.get('local'),
            "artigo": request.form.get('enquadramento'),
            "testemunhas": request.form.get('testemunha'),
            "comunicante": comunicante,
            "observacoes": request.form.get('observacoes'),
        }
                
        response = requests.post('http://127.0.0.1:8000/cpi/create', json=cpi, headers=headers)
        
        
        if request.form.get('submit') == 'novo':
            return redirect(url_for('cpi.adicionar_cpi'))
        return redirect(url_for('usuario.home'))


@cpi.route('/perfil_cpi', methods=['GET'])
def perfil_cpi():
    if not session.get('token'):
        return redirect(url_for('usuario.login'))
    
    headers = {'Authorization': f'Bearer {session["token"]}'}
    
    response = requests.get(f'http://127.0.0.1:8000/usuario/get_usuario', headers=headers)
    
    if not response:
        del session['token']
        return redirect(url_for('usuario.login', erro=1))
    
    usuario = response.json()['__data__']
    
    
    protocolo = request.args.get('protocolo', type=int)
    if not protocolo:
        return redirect(url_for('consulta.consultar'))
    
    response = requests.get(f'http://127.0.0.1:8000/cpi/get_by_cpi_id?id={protocolo}', headers=headers)
    
    if not response:
        return redirect(url_for('consulta.consultar', erro=1))
    else:
        cpi = response.json()
        
        response = requests.get(f'http://127.0.0.1:8000/defesa/get_by_cpi', params={'cpi_id': cpi['id']}, headers=headers)
        if response:
            defesa = response.json()
        else:
            defesa = None
                                
        response = requests.get(f'http://127.0.0.1:8000/parecer_cmd_cia/get_by_cpi', params={'cpi_id': cpi['id']}, headers=headers)
        if response:
            parecer_cmd_cia = response.json()
        else:
            parecer_cmd_cia = None
                                
        response = requests.get(f'http://127.0.0.1:8000/parecer/get_by_cpi', params={'cpi_id': cpi['id']}, headers=headers)
        if response:
            parecer = response.json()
        else:
            parecer = None
        
                        
        response = requests.get(f'http://127.0.0.1:8000/decisao/get_by_cpi', params={'cpi_id': cpi['id']}, headers=headers)
        if response:
            decisao = response.json()
        else:
            decisao = None
                        
        return render_template('cpi.html', cpi=cpi, defesa=defesa, parecer=parecer, parecer_cmd_cia=parecer_cmd_cia, decisao=decisao)
    
    
@cpi.route('/consultar_caderno', methods=['GET', 'POST'])
def consultar_caderno():
    if not session.get('token'):
        return redirect(url_for('usuario.login'))
    
    headers = {'Authorization': f'Bearer {session["token"]}'}
    
    response = requests.get(f'http://127.0.0.1:8000/usuario/get_usuario', headers=headers)
    
    if not response:
        del session['token']
        return redirect(url_for('usuario.login', erro=1))
    
    usuario = response.json()['__data__']
    
    if request.method == 'GET':
        return render_template('consultar_caderno.html', cadernos=None, usuario=usuario)
    
    tipo = request.form.get('tipo')
    
    temp = request.form.get('data_inicial')
    temp = temp.split('-')
    data_inicial = f'{temp[2]}/{temp[1]}/{temp[0]}'
    
    temp = request.form.get('data_final')  
    temp = temp.split('-')
    data_final = f'{temp[2]}/{temp[1]}/{temp[0]}'
      
    pelotoes = request.form.get('pelotoes', type=int)  
    conduta = request.form.get('conduta')  
    
    curso = request.form.get('curso')
    ano = request.form.get('ano')
    
    if tipo == 'consultar':
    
        response = requests.get(f'http://127.0.0.1:8000/cpi/consulta_data?data_inicial={data_inicial}&data_final={data_final}&conduta={conduta}&curso={curso}&ano={ano}', headers=headers)
        
        if not response:
            return redirect(url_for('cpi.consultar_caderno', erro=1))
            
        cadernos = response.json()
        
        quantidade_aluno_cpi1 = 0
        quantidade_aluno_cpi2 = 0
        quantidade_aluno_cpi3 = 0
        
        for caderno in cadernos:
            conduta = caderno['conduta']
            if conduta == 3:
                quantidade_aluno_cpi3 += 1
            elif conduta == 2:
                quantidade_aluno_cpi2 += 1
            elif conduta == 1:
                quantidade_aluno_cpi1 += 1
        
        conduta = request.form.get('conduta').upper()
        
        return render_template('consultar_caderno.html', cadernos=cadernos, usuario=usuario, pelotoes=pelotoes, quantidade_aluno_cpi1=quantidade_aluno_cpi1, quantidade_aluno_cpi2=quantidade_aluno_cpi2, quantidade_aluno_cpi3=quantidade_aluno_cpi3, conduta=conduta)

    elif tipo == 'gerar':        
        response = requests.get(f'http://127.0.0.1:8000/cpi/gerar_caderno?data_inicial={data_inicial}&data_final={data_final}&curso={curso}&ano={ano}', headers=headers)
        
        if not response:
            return redirect(url_for('cpi.consultar_caderno', erro=1))
        
        cadernos = response.json()
        
        quantidade_aluno_cpi1 = 0
        quantidade_aluno_cpi2 = 0
        quantidade_aluno_cpi3 = 0
        
        for caderno in cadernos['caderno_cpi']:
            conduta = caderno['conduta']
            if conduta == 3:
                quantidade_aluno_cpi3 += 1
            elif conduta == 2:
                quantidade_aluno_cpi2 += 1
            elif conduta == 1:
                quantidade_aluno_cpi1 += 1
                
        total_cpa = cadernos['total_cpa']
        caderno_cpi = cadernos['caderno_cpi']
        caderno_cpa = cadernos['caderno_cpa']
        
        #####  GERANDO PLANILHA  #####
        
        FILE_NAME = f'{date.today()}_{usuario["nome_de_guerra"]}.xlsx'
        y = 1
        
        A, B, C, D, E = 1, 2, 3, 4, 5
        
        workbook = Workbook()
        sheet_cpi = workbook.active
        sheet_cpi.title = 'CADERNO CPI'
        
                
        sheet_cpi.merge_cells(f'A{y}:E{y}')
        
        font_titulo = Font(bold=True, size=12)
        alignment = Alignment('center', 'center')
        side = Side('thin', color='000000')
        border = Border(side, side, side, side)
        fill_cpi1 = PatternFill('solid', 'ffff00')
        fill_cpi2 = PatternFill('solid', 'ff9b1f')
        fill_cpi3 = PatternFill('solid', 'ff0000')
        
        sheet_cpi['G1'] = 'TOTAL CPI I'
        sheet_cpi['G2'] = 'TOTAL CPI II'
        sheet_cpi['G3'] = 'TOTAL CPI III'
        sheet_cpi['G5'] = 'TOTAL CPI'
        
        sheet_cpi['H1'] = quantidade_aluno_cpi1
        sheet_cpi['H2'] = quantidade_aluno_cpi2
        sheet_cpi['H3'] = quantidade_aluno_cpi3
        sheet_cpi['H5'] = quantidade_aluno_cpi1 + quantidade_aluno_cpi2 + quantidade_aluno_cpi3        
        
        
        sheet_cpi.column_dimensions['A'].width = 10
        sheet_cpi.column_dimensions['B'].width = 20
        sheet_cpi.column_dimensions['C'].width = 15
        sheet_cpi.column_dimensions['D'].width = 25
        sheet_cpi.column_dimensions['E'].width = 20
        sheet_cpi.column_dimensions['G'].width = 20
        sheet_cpi.column_dimensions['H'].width = 10
        
        
        sheet_cpi[f'A{y}'] = 'CPI I'
        sheet_cpi.cell(y, A).font = font_titulo
        sheet_cpi.cell(y, A).alignment = alignment
        sheet_cpi.cell(y, A).fill = fill_cpi1
        sheet_cpi.cell(y, A).border = border
        sheet_cpi.cell(y, B).border = border
        sheet_cpi.cell(y, C).border = border
        sheet_cpi.cell(y, D).border = border
        sheet_cpi.cell(y, E).border = border
        
        y += 1
        
        sheet_cpi[f'A{y}'] = 'Nº'
        if curso.upper() == 'CFSD':
            sheet_cpi[f'B{y}'] = 'AL SD'
        elif curso.upper() == 'CFO':
            sheet_cpi[f'B{y}'] = 'AL OF'
        elif curso.upper() == 'CHS':
            sheet_cpi[f'B{y}'] = 'AL SGT'
        sheet_cpi[f'C{y}'] = 'DATA'
        sheet_cpi[f'D{y}'] = 'ENQUADRAMENTO'
        sheet_cpi[f'E{y}'] = 'PROTOCOLO'
        
        sheet_cpi.cell(y, A).font = font_titulo
        sheet_cpi.cell(y, A).alignment = alignment
        sheet_cpi.cell(y, A).border = border
        
        sheet_cpi.cell(y, B).font = font_titulo
        sheet_cpi.cell(y, B).alignment = alignment
        sheet_cpi.cell(y, B).border = border
        
        sheet_cpi.cell(y, C).font = font_titulo
        sheet_cpi.cell(y, C).alignment = alignment
        sheet_cpi.cell(y, C).border = border
        
        sheet_cpi.cell(y, D).font = font_titulo
        sheet_cpi.cell(y, D).alignment = alignment
        sheet_cpi.cell(y, D).border = border
        
        sheet_cpi.cell(y, E).font = font_titulo
        sheet_cpi.cell(y, E).alignment = alignment
        sheet_cpi.cell(y, E).border = border
        
        y += 1
        
        for i in range(1, pelotoes+1):
            sheet_cpi.merge_cells(f'A{y}:E{y}')
            sheet_cpi[f'A{y}'] = f'{i}º PELOTÃO'
            
            sheet_cpi.cell(y, A).font = font_titulo
            sheet_cpi.cell(y, A).alignment = alignment
            sheet_cpi.cell(y, A).fill = fill_cpi1
            sheet_cpi.cell(y, A).border = border
            sheet_cpi.cell(y, B).border = border
            sheet_cpi.cell(y, C).border = border
            sheet_cpi.cell(y, D).border = border
            sheet_cpi.cell(y, E).border = border
            
            y += 1
            
            sheet_cpi.cell(y, A).alignment = alignment
            sheet_cpi.cell(y, B).alignment = alignment
            sheet_cpi.cell(y, C).alignment = alignment
            sheet_cpi.cell(y, D).alignment = alignment
            sheet_cpi.cell(y, E).alignment = alignment
            
            sheet_cpi.cell(y, A).border = border
            sheet_cpi.cell(y, B).border = border
            sheet_cpi.cell(y, C).border = border
            sheet_cpi.cell(y, D).border = border
            sheet_cpi.cell(y, E).border = border
            
            count = False
            
            for caderno in caderno_cpi:
                if caderno['conduta'] == 1 and int(caderno['info']['pelotao']) == i:
                    count = True
                    sheet_cpi.merge_cells(f'A{y}:A{y+len(caderno["cpis"])-1}')
                    sheet_cpi.merge_cells(f'B{y}:B{y+len(caderno["cpis"])-1}')
                    
                    sheet_cpi[f'A{y}'] = caderno['info']['numero']
                    sheet_cpi[f'B{y}'] = caderno['aluno']['nome_de_guerra']
                    
                    for cpi in caderno["cpis"]:
                        sheet_cpi[f'C{y}'] = cpi['data']
                        sheet_cpi[f'D{y}'] = cpi['artigo']
                        sheet_cpi[f'E{y}'] = cpi['id']
                        
                        y += 1
                        
                        sheet_cpi.cell(y, A).alignment = alignment
                        sheet_cpi.cell(y, B).alignment = alignment
                        sheet_cpi.cell(y, C).alignment = alignment
                        sheet_cpi.cell(y, D).alignment = alignment
                        sheet_cpi.cell(y, E).alignment = alignment
                        
                        sheet_cpi.cell(y, A).border = border
                        sheet_cpi.cell(y, B).border = border
                        sheet_cpi.cell(y, C).border = border
                        sheet_cpi.cell(y, D).border = border
                        sheet_cpi.cell(y, E).border = border
            
            
            if not count:
                sheet_cpi.merge_cells(f'A{y}:E{y}')
                sheet_cpi[f'A{y}'] = '...'
                sheet_cpi.cell(y, A).border = border
                sheet_cpi.cell(y, B).border = border
                sheet_cpi.cell(y, C).border = border
                sheet_cpi.cell(y, D).border = border
                sheet_cpi.cell(y, E).border = border
                                
                y += 1
                
        # CPI II
        
        y += 1
        
        
        sheet_cpi.merge_cells(f'A{y}:E{y}')
        sheet_cpi[f'A{y}'] = 'CPI II'
        sheet_cpi.cell(y, A).font = font_titulo
        sheet_cpi.cell(y, A).alignment = alignment
        sheet_cpi.cell(y, A).fill = fill_cpi2
        sheet_cpi.cell(y, A).border = border
        sheet_cpi.cell(y, B).border = border
        sheet_cpi.cell(y, C).border = border
        sheet_cpi.cell(y, D).border = border
        sheet_cpi.cell(y, E).border = border
        
        y += 1
        
        sheet_cpi[f'A{y}'] = 'Nº'
        if curso.upper() == 'CFSD':
            sheet_cpi[f'B{y}'] = 'AL SD'
        elif curso.upper() == 'CFO':
            sheet_cpi[f'B{y}'] = 'AL OF'
        elif curso.upper() == 'CHS':
            sheet_cpi[f'B{y}'] = 'AL SGT'
        sheet_cpi[f'C{y}'] = 'DATA'
        sheet_cpi[f'D{y}'] = 'ENQUADRAMENTO'
        sheet_cpi[f'E{y}'] = 'PROTOCOLO'
        
        sheet_cpi.cell(y, A).font = font_titulo
        sheet_cpi.cell(y, A).alignment = alignment
        sheet_cpi.cell(y, A).border = border
        
        sheet_cpi.cell(y, B).font = font_titulo
        sheet_cpi.cell(y, B).alignment = alignment
        sheet_cpi.cell(y, B).border = border

        sheet_cpi.cell(y, C).font = font_titulo
        sheet_cpi.cell(y, C).alignment = alignment
        sheet_cpi.cell(y, C).border = border
        
        sheet_cpi.cell(y, D).font = font_titulo
        sheet_cpi.cell(y, D).alignment = alignment
        sheet_cpi.cell(y, D).border = border
        
        sheet_cpi.cell(y, E).font = font_titulo
        sheet_cpi.cell(y, E).alignment = alignment
        sheet_cpi.cell(y, E).border = border
        
        y += 1
        
        for i in range(1, pelotoes+1):
            sheet_cpi.merge_cells(f'A{y}:E{y}')
            sheet_cpi[f'A{y}'] = f'{i}º PELOTÃO'
            
            sheet_cpi.cell(y, A).font = font_titulo
            sheet_cpi.cell(y, A).alignment = alignment
            sheet_cpi.cell(y, A).fill = fill_cpi2
            sheet_cpi.cell(y, A).border = border
            sheet_cpi.cell(y, B).border = border
            sheet_cpi.cell(y, C).border = border
            sheet_cpi.cell(y, D).border = border
            sheet_cpi.cell(y, E).border = border
            
            y += 1
            
            count = False
            
            sheet_cpi.cell(y, A).alignment = alignment
            sheet_cpi.cell(y, B).alignment = alignment
            sheet_cpi.cell(y, C).alignment = alignment
            sheet_cpi.cell(y, D).alignment = alignment
            sheet_cpi.cell(y, E).alignment = alignment
            
            sheet_cpi.cell(y, A).border = border
            sheet_cpi.cell(y, B).border = border
            sheet_cpi.cell(y, C).border = border
            sheet_cpi.cell(y, D).border = border
            sheet_cpi.cell(y, E).border = border
            
            for caderno in caderno_cpi:
                if caderno['conduta'] == 2 and int(caderno['info']['pelotao']) == i:
                    count = True
                    sheet_cpi.merge_cells(f'A{y}:A{y+len(caderno["cpis"])-1}')
                    sheet_cpi.merge_cells(f'B{y}:B{y+len(caderno["cpis"])-1}')
                    
                    sheet_cpi[f'A{y}'] = caderno['info']['numero']
                    sheet_cpi[f'B{y}'] = caderno['aluno']['nome_de_guerra']
                    
                    for cpi in caderno["cpis"]:
                        sheet_cpi[f'C{y}'] = cpi['data']
                        sheet_cpi[f'D{y}'] = cpi['artigo']
                        sheet_cpi[f'E{y}'] = cpi['id']
                        
                        y += 1
                        
                        sheet_cpi.cell(y, A).alignment = alignment
                        sheet_cpi.cell(y, B).alignment = alignment
                        sheet_cpi.cell(y, C).alignment = alignment
                        sheet_cpi.cell(y, D).alignment = alignment
                        sheet_cpi.cell(y, E).alignment = alignment
                        
                        sheet_cpi.cell(y, A).border = border
                        sheet_cpi.cell(y, B).border = border
                        sheet_cpi.cell(y, C).border = border
                        sheet_cpi.cell(y, D).border = border
                        sheet_cpi.cell(y, E).border = border
            
            if not count:
                sheet_cpi.merge_cells(f'A{y}:E{y}')
                sheet_cpi[f'A{y}'] = '...'
                sheet_cpi.cell(y, A).alignment = alignment
                sheet_cpi.cell(y, A).border = border
                sheet_cpi.cell(y, B).border = border
                sheet_cpi.cell(y, C).border = border
                sheet_cpi.cell(y, D).border = border
                sheet_cpi.cell(y, E).border = border
                
                y += 1
                
                        
        # CPI III
        y += 1
        
        sheet_cpi.merge_cells(f'A{y}:E{y}')
        sheet_cpi[f'A{y}'] = 'CPI III'
        sheet_cpi.cell(y, A).font = font_titulo
        sheet_cpi.cell(y, A).alignment = alignment
        sheet_cpi.cell(y, A).fill = fill_cpi3
        sheet_cpi.cell(y, A).border = border
        sheet_cpi.cell(y, B).border = border
        sheet_cpi.cell(y, C).border = border
        sheet_cpi.cell(y, D).border = border
        sheet_cpi.cell(y, E).border = border
        
        y += 1
        
        sheet_cpi[f'A{y}'] = 'Nº'
        if curso.upper() == 'CFSD':
            sheet_cpi[f'B{y}'] = 'AL SD'
        elif curso.upper() == 'CFO':
            sheet_cpi[f'B{y}'] = 'AL OF'
        elif curso.upper() == 'CHS':
            sheet_cpi[f'B{y}'] = 'AL SGT'
        sheet_cpi[f'C{y}'] = 'DATA'
        sheet_cpi[f'D{y}'] = 'ENQUADRAMENTO'
        sheet_cpi[f'E{y}'] = 'PROTOCOLO'
        
        sheet_cpi.cell(y, A).font = font_titulo
        sheet_cpi.cell(y, A).alignment = alignment
        sheet_cpi.cell(y, A).border = border
        
        sheet_cpi.cell(y, B).font = font_titulo
        sheet_cpi.cell(y, B).alignment = alignment
        sheet_cpi.cell(y, B).border = border
        
        sheet_cpi.cell(y, C).font = font_titulo
        sheet_cpi.cell(y, C).alignment = alignment
        sheet_cpi.cell(y, C).border = border
        
        sheet_cpi.cell(y, D).font = font_titulo
        sheet_cpi.cell(y, D).alignment = alignment
        sheet_cpi.cell(y, D).border = border
        
        sheet_cpi.cell(y, E).font = font_titulo
        sheet_cpi.cell(y, E).alignment = alignment
        sheet_cpi.cell(y, E).border = border
        
        y += 1
        
        sheet_cpi.cell(y, A).alignment = alignment
        sheet_cpi.cell(y, B).alignment = alignment
        sheet_cpi.cell(y, C).alignment = alignment
        sheet_cpi.cell(y, D).alignment = alignment
        sheet_cpi.cell(y, E).alignment = alignment
        
        sheet_cpi.cell(y, A).border = border
        sheet_cpi.cell(y, B).border = border
        sheet_cpi.cell(y, C).border = border
        sheet_cpi.cell(y, D).border = border
        sheet_cpi.cell(y, E).border = border
        
        for i in range(1, pelotoes+1):
            sheet_cpi.merge_cells(f'A{y}:E{y}')
            sheet_cpi[f'A{y}'] = f'{i}º PELOTÃO'
            
            sheet_cpi.cell(y, A).font = font_titulo
            sheet_cpi.cell(y, A).alignment = alignment
            sheet_cpi.cell(y, A).fill = fill_cpi3
            sheet_cpi.cell(y, A).border = border
            
            y += 1
            
            count = False
            
            sheet_cpi.cell(y, A).alignment = alignment
            sheet_cpi.cell(y, B).alignment = alignment
            sheet_cpi.cell(y, C).alignment = alignment
            sheet_cpi.cell(y, D).alignment = alignment
            sheet_cpi.cell(y, E).alignment = alignment
            
            sheet_cpi.cell(y, A).border = border
            sheet_cpi.cell(y, B).border = border
            sheet_cpi.cell(y, C).border = border
            sheet_cpi.cell(y, D).border = border
            sheet_cpi.cell(y, E).border = border
            
            for caderno in caderno_cpi:
                if caderno['conduta'] == 3 and int(caderno['info']['pelotao']) == i:
                    count = True
                    sheet_cpi.merge_cells(f'A{y}:A{y+len(caderno["cpis"])-1}')
                    sheet_cpi.merge_cells(f'B{y}:B{y+len(caderno["cpis"])-1}')
                    
                    sheet_cpi[f'A{y}'] = caderno['info']['numero']
                    sheet_cpi[f'B{y}'] = caderno['aluno']['nome_de_guerra']
                    
                    for cpi in caderno["cpis"]:
                        sheet_cpi[f'C{y}'] = cpi['data']
                        sheet_cpi[f'D{y}'] = cpi['artigo']
                        sheet_cpi[f'E{y}'] = cpi['id']
                        
                        y += 1
                        
                        sheet_cpi.cell(y, A).alignment = alignment
                        sheet_cpi.cell(y, B).alignment = alignment
                        sheet_cpi.cell(y, C).alignment = alignment
                        sheet_cpi.cell(y, D).alignment = alignment
                        sheet_cpi.cell(y, E).alignment = alignment
                        
                        sheet_cpi.cell(y, A).border = border
                        sheet_cpi.cell(y, B).border = border
                        sheet_cpi.cell(y, C).border = border
                        sheet_cpi.cell(y, D).border = border
                        sheet_cpi.cell(y, E).border = border
            
            
            if not count:
                sheet_cpi.merge_cells(f'A{y}:E{y}')
                sheet_cpi[f'A{y}'] = '...'
                sheet_cpi.cell(y, A).alignment = alignment
                sheet_cpi.cell(y, A).border = border
                sheet_cpi.cell(y, B).border = border
                sheet_cpi.cell(y, C).border = border
                sheet_cpi.cell(y, D).border = border
                sheet_cpi.cell(y, E).border = border
                
                y += 1
                
        #### SHEET CADERNO CPA ####
        
        sheet_cpa = workbook.create_sheet('CADERNO CPA')
        y = 1
        
        sheet_cpa.merge_cells(f'A{y}:E{y}')
        
        font_titulo = Font(bold=True, size=12)
        alignment = Alignment('center', 'center')
        side = Side('thin', color='000000')
        border = Border(side, side, side, side)
        fill_cpi1 = PatternFill('solid', 'ffffff')
        fill_cpi2 = PatternFill('solid', 'ffffff')
        fill_cpi3 = PatternFill('solid', 'ffffff')
        
        sheet_cpa['G2'] = 'TOTAL CPA'
        
        sheet_cpa['H2'] = total_cpa
        
        sheet_cpa.column_dimensions['A'].width = 10
        sheet_cpa.column_dimensions['B'].width = 20
        sheet_cpa.column_dimensions['C'].width = 15
        sheet_cpa.column_dimensions['D'].width = 25
        sheet_cpa.column_dimensions['E'].width = 20
        sheet_cpa.column_dimensions['G'].width = 20
        sheet_cpa.column_dimensions['H'].width = 10
        
        sheet_cpa[f'A{y}'] = 'REFERÊNCIAS ELOGIOSAS'
        sheet_cpa.cell(y, A).font = font_titulo
        sheet_cpa.cell(y, A).alignment = alignment
        sheet_cpa.cell(y, A).fill = fill_cpi1
        sheet_cpa.cell(y, A).border = border
        sheet_cpa.cell(y, B).border = border
        sheet_cpa.cell(y, C).border = border
        sheet_cpa.cell(y, D).border = border
        sheet_cpa.cell(y, E).border = border
        
        y += 1
        
        sheet_cpa[f'A{y}'] = 'Nº'
        if curso.upper() == 'CFSD':
            sheet_cpa[f'B{y}'] = 'AL SD'
        elif curso.upper() == 'CFO':
            sheet_cpa[f'B{y}'] = 'AL OF'
        elif curso.upper() == 'CHS':
            sheet_cpa[f'B{y}'] = 'AL SGT'
        sheet_cpa[f'C{y}'] = 'DATA'
        sheet_cpa[f'D{y}'] = 'ENQUADRAMENTO'
        sheet_cpa[f'E{y}'] = 'PROTOCOLO'
        
        sheet_cpa.cell(y, A).font = font_titulo
        sheet_cpa.cell(y, A).alignment = alignment
        sheet_cpa.cell(y, A).border = border
        
        sheet_cpa.cell(y, B).font = font_titulo
        sheet_cpa.cell(y, B).alignment = alignment
        sheet_cpa.cell(y, B).border = border
        
        sheet_cpa.cell(y, C).font = font_titulo
        sheet_cpa.cell(y, C).alignment = alignment
        sheet_cpa.cell(y, C).border = border
        
        sheet_cpa.cell(y, D).font = font_titulo
        sheet_cpa.cell(y, D).alignment = alignment
        sheet_cpa.cell(y, D).border = border
        
        sheet_cpa.cell(y, E).font = font_titulo
        sheet_cpa.cell(y, E).alignment = alignment
        sheet_cpa.cell(y, E).border = border
        
        y += 1
        
        for i in range(1, pelotoes+1):
            sheet_cpa.merge_cells(f'A{y}:E{y}')
            sheet_cpa[f'A{y}'] = f'{i}º PELOTÃO'
            
            sheet_cpa.cell(y, A).font = font_titulo
            sheet_cpa.cell(y, A).alignment = alignment
            sheet_cpa.cell(y, A).fill = fill_cpi1
            sheet_cpa.cell(y, A).border = border
            sheet_cpa.cell(y, B).border = border
            sheet_cpa.cell(y, C).border = border
            sheet_cpa.cell(y, D).border = border
            sheet_cpa.cell(y, E).border = border
            
            y += 1
            
            sheet_cpa.cell(y, A).alignment = alignment
            sheet_cpa.cell(y, B).alignment = alignment
            sheet_cpa.cell(y, C).alignment = alignment
            sheet_cpa.cell(y, D).alignment = alignment
            sheet_cpa.cell(y, E).alignment = alignment
            
            sheet_cpa.cell(y, A).border = border
            sheet_cpa.cell(y, B).border = border
            sheet_cpa.cell(y, C).border = border
            sheet_cpa.cell(y, D).border = border
            sheet_cpa.cell(y, E).border = border
            
            count = False
            
            for caderno in caderno_cpa:
                if caderno['conduta'] == 1 and int(caderno['info']['pelotao']) == i:
                    count = True
                    sheet_cpa.merge_cells(f'A{y}:A{y+len(caderno["cpis"])-1}')
                    sheet_cpa.merge_cells(f'B{y}:B{y+len(caderno["cpis"])-1}')
                    
                    sheet_cpa[f'A{y}'] = caderno['info']['numero']
                    sheet_cpa[f'B{y}'] = caderno['aluno']['nome_de_guerra']
                    
                    for cpi in caderno["cpis"]:
                        sheet_cpa[f'C{y}'] = cpi['data']
                        sheet_cpa[f'D{y}'] = cpi['artigo']
                        sheet_cpa[f'E{y}'] = cpi['id']
                        
                        y += 1
                        
                        sheet_cpa.cell(y, A).alignment = alignment
                        sheet_cpa.cell(y, B).alignment = alignment
                        sheet_cpa.cell(y, C).alignment = alignment
                        sheet_cpa.cell(y, D).alignment = alignment
                        sheet_cpa.cell(y, E).alignment = alignment
                        
                        sheet_cpa.cell(y, A).border = border
                        sheet_cpa.cell(y, B).border = border
                        sheet_cpa.cell(y, C).border = border
                        sheet_cpa.cell(y, D).border = border
                        sheet_cpa.cell(y, E).border = border
            
            
            if not count:
                sheet_cpa.merge_cells(f'A{y}:E{y}')
                sheet_cpa[f'A{y}'] = '...'
                sheet_cpa.cell(y, A).border = border
                sheet_cpa.cell(y, B).border = border
                sheet_cpa.cell(y, C).border = border
                sheet_cpa.cell(y, D).border = border
                sheet_cpa.cell(y, E).border = border
                                
                y += 1
        
        
        workbook.save(f'{FILE_NAME}')
                
        
        return send_file(f'{FILE_NAME}')
        # return render_template('gerar_caderno.html', caderno_cpi=caderno_cpi, caderno_cpa=caderno_cpa, total_cpa=total_cpa, usuario=usuario, pelotoes=pelotoes, quantidade_aluno_cpi1=quantidade_aluno_cpi1, quantidade_aluno_cpi2=quantidade_aluno_cpi2, quantidade_aluno_cpi3=quantidade_aluno_cpi3)